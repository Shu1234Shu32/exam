import requests
import openpyxl
from openpyxl import Workbook

def get_distance(api_key, origin, destination, travel_mode='driving'):
    base_url = "https://dev.virtualearth.net/REST/v1/Routes/"
    params = {
        'key': api_key,
        'wayPoint.1': origin,
        'wayPoint.2': destination,
        'optimize': 'distance',
        'routeAttributes': 'routePath',
        'travelMode': travel_mode,
        'distanceUnit':"km"
    }
    response = requests.get(base_url + travel_mode, params=params)
    if response.status_code == 200:
        data = response.json()
        if 'resourceSets' in data and data['resourceSets'] and 'resources' in data['resourceSets'][0]:
            route = data['resourceSets'][0]['resources'][0]
            distance = route['travelDistance']
            return distance
        else:
            return None
    else:
        print(f"Error: {response.status_code}")
        return None

if __name__ == "__main__":
    bing_maps_api_key = 'ArJWc58bcqR4mNklHYm7IXeeaJtAj4imCjuEKfV6cdAIeDV3ruronaHgc6WlPMuP'
    input_file = "/path/to/root/directory"
    output_file = "output.xlsx"
    
    wb = openpyxl.load_workbook(input_file, data_only=True)
    sheet = wb['Input']
    max_row = sheet.max_row
    
    for row_number in range(1, max_row + 1):
        origin_address = sheet.cell(row=row_number, column=2).value
        destination_address = sheet.cell(row=row_number, column=3).value
        travel_mode = 'Driving'
        distance = get_distance(bing_maps_api_key, origin_address, destination_address, travel_mode)
        if distance is not None:
            print(f"Distance from {origin_address} to {destination_address}: {distance} kms")
            sheet.cell(row=row_number, column=4).value = distance
