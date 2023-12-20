import requests
import openpyxl
from openpyxl import Workbook

port={
      "Unites States": "port-of-new-york,united-states",
      "China": "port-of-shanghai,china",
      "India": "port-of-mumbai,india",
      "Dubai": "port-of-mina-jabal-ali-jebel-ali,united-arab-emirates",
      "United Kingdom": "port-of-felixstowe,united-kingdom"
      }

def get_country(api_key, location):
    base_url = "http://dev.virtualearth.net/REST/v1/Locations"
    params = {
        'query': location,
        'key': api_key
    }
    response = requests.get(base_url, params=params)
    if response.status_code == 200:
        data = response.json()
        resource_sets = data.get('resourceSets', [])

        if resource_sets:
            resources = resource_sets[0].get('resources', [])

            if resources:
                address = resources[0].get('address', {})
                country = address.get('countryRegion')
                return country
    print(f"Error: Unable to retrieve country for {location}")
    return None


def get_distance(api_key, origin, destination, travel_mode='driving'):
    base_url = "https://dev.virtualearth.net/REST/v1/Routes/"
    params = {
        'key': api_key,
        'wayPoint.1': origin,
        'wayPoint.2': destination,
        'optimize': 'distance',
        'routeAttributes': 'routePath',
        'travelMode': travel_mode,
        'distanceUnit':"km"
    }
    response = requests.get(base_url + travel_mode, params=params)
    if response.status_code == 200:
        data = response.json()
        if 'resourceSets' in data and data['resourceSets'] and 'resources' in data['resourceSets'][0]:
            route = data['resourceSets'][0]['resources'][0]
            distance = route['travelDistance']
            return distance
        else:
            return None
    else:
        print(f"Error: {response.status_code}")
        return None

if __name__ == "__main__":
    bing_maps_api_key = 'ArJWc58bcqR4mNklHYm7IXeeaJtAj4imCjuEKfV6cdAIeDV3ruronaHgc6WlPMuP'
    #input_file = "/path/to/root/directory"
    #output_file = "output.xlsx"
    
    wb = openpyxl.load_workbook(input_file, data_only=True)
    #sheet = wb['Input']
    max_row = sheet.max_row
    sheet.cell(row=1, column=4).value = "Road Distance"
    sheet.cell(row=1, column=5).value = "Trans Oceanic Distance"
    sheet.cell(row=1, column=6).value = "Road Distance"
    
    for row_number in range(1, max_row + 1):
        #origin_address = sheet.cell(row=row_number, column=2).value
        #destination_address = sheet.cell(row=row_number, column=3).value
        travel_mode = 'Driving'
        distance = get_distance(bing_maps_api_key, origin_address, destination_address, travel_mode)
        if distance is not None:
            print(f"Distance from {origin_address} to {destination_address}: {distance} kms")
            sheet.cell(row=row_number, column=4).value = distance
        else:
            country_1 = get_country(bing_maps_api_key, source)
            country_2 = get_country(bing_maps_api_key, destination)
            
            distance = get_distance(bing_maps_api_key, origin_address, port[country_1], travel_mode)
            if distance is not None:
                print(f"Distance from {origin_address} to {port[country_1]}: {distance} kms")
                sheet.cell(row=row_number, column=4).value = distance
            
            response = requests.get("http://ports.com/sea-route/ardmair,united-kingdom/port-of-hanoi,vietnam/")
            res = response.text.split(" ")
            distance = res[res.index('nautical')-1]
            sheet.cell(row=row_number, column=5).value = distance
            print(f"Distance from {port[country_1]} to {port[country_2]}: {distance} kms")
            
            distance = get_distance(bing_maps_api_key, port[country_2], destination_address, travel_mode)
            if distance is not None:
                print(f"Distance from {port[country_2]} to {destination_address}: {distance} kms")
                sheet.cell(row=row_number, column=6).value = distance
            